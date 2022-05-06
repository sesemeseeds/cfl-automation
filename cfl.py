from __future__ import print_function

import email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import os.path

from bs4 import BeautifulSoup
import openpyxl

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import base64

from pyparsing import col


# If modifying these scopes, delete the file token.json
SCOPES = ['https://mail.google.com/', 'https://www.googleapis.com/auth/drive']

# excel info file path, excel is in root folder, so no need to change
SPREADSHEET_PATH = "customer-info.xlsx"

#owner email
MY_EMAIL = "cincinnatifilmlab@gmail.com"    

"""
MAKE SURE EXCEL FILE IS CLOSED AND SAVED
"""

def main():
    gmail_service, drive_service = google_service()             # fetches service auth from Google gmail and drive service

    workbook = openpyxl.load_workbook(SPREADSHEET_PATH)         # opens excel file
    print("Opening workbook... SUCCESS\n")
    ws0 = workbook.active                                  # assigns workbook to variable named ws
    ws = ws0['Order']

    progressCol = ws['D']                                       # looks at column D (progress column) ['In progress', 'Ready', 'Done']
    orderCol = ws['A']
    
    for cell in progressCol:                                    # for each cell...
        if (cell.value == "Ready"):                             # finds which cells are 'Ready'
            cellRow = cell.row                                  # stores row number to find instance

            # VARIABLE ASSIGNMENT
            orderNum = ws['A' + str(cellRow)].value             # stores orderNum
            print("Finding \'Ready\' order...SUCCESS (Order #" + str(orderNum) + ")\n")
            name = ws['B' + str(cellRow)].value                 # stores order name
            email = ws['C' + str(cellRow)].value                # stores order email

            #RETRIEVES SHARE LINK
            share_link = str(find_folder(drive_service, str(orderNum)))     

            # DRIVE FOLDER DOES NOT EXIST: this happens when there is a order in the excel file marked as ready, but there is no folder in the Drive
            if (share_link == "None"):                     
                print("ERROR: No photos folder was found for Order #" + str(orderNum) + ". Please check if order folder exists within Drive\n")
                continue
            else:        
                print("Finding share link for Order #"+ str(orderNum) + "... SUCCESS\n")

                # FORMATTING THE EMAIL
                subject = "Cincinnati Film Lab has your photos ready! Order #" + str(orderNum)              
                body = "Hi " + name + "! Your photos have been processed and are available at the link below. Thank you! \n\n"+ share_link + "\n\nTo pick up your negatives, schedule a time here: https://app.squarespacescheduling.com/schedule.php?owner=23693339."
                send_message(gmail_service, email, subject, body)

                for cell in orderCol:
                    if(cell.value == orderNum):
                        r = cell.row
                        ws['D' + str(r)] = "Done"

                workbook.save(SPREADSHEET_PATH)



def google_service():
    """
    Shows basic usage of the Gmail API and Google Drive API    
    Lists the user's Gmail labels.
    """
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=52150)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    try:                     
        # Call the Gmail API
        GMAILservice = build('gmail', 'v1', credentials=creds)

        # Call the Drive v3 API
        DRIVEservice = build('drive', 'v3', credentials=creds)

        print("\nCalling Google Services API... SUCCESS\n")

        return GMAILservice, DRIVEservice

    except HttpError as error:
        # TODO(developer) - Handle errors from gmail API.
        print(f'An error occurred: {error}')


def find_folder(service, order) :
    """Finds the order folder.

    Args:
        service: generated drive auth service in google_service()
        order: order number

    Returns:
        A weblink to the folder to share with the customer
    """
    results = service.files().list(pageSize=10, fields="nextPageToken, files(id, name)").execute()                  # API retrieves all files and folders of Drive
    items = results.get('files', [])                                                                                
    order = "Order #" + order                   # reformats orderNum 
    
    request_body = {                            # json of request permissions
        "role": "reader",                       
        "type": "anyone"        
    }

    for item in items:                          # for each item in the Google Drive
        if item['name'] == order:               # if the name of the folder matches the orderNum    
            file_id = item['id']                # file_id is stored (folder id)

            service.permissions().create(fileId = file_id, body = request_body).execute()           # permissions of the folder is changed

            response_share_link = service.files().get(fileId = file_id, fields='webViewLink').execute()         # share link of the folder is stored...
            return response_share_link["webViewLink"]                                                           # and returned

def send_message(service, to, subject, message_text):
    """Create a message for an email.

    Args:
    sender: Email address of the sender.
    to: Email address of the receiver.
    subject: The subject of the email message.
    message_text: The text of the email message.

    Returns:
    An object containing a base64url encoded email object.
    """
    mimeMessage = MIMEMultipart()
    mimeMessage['to'] = to
    mimeMessage['subject'] = subject
    mimeMessage.attach(MIMEText(message_text, 'plain'))
    raw_string = base64.urlsafe_b64encode(mimeMessage.as_bytes()).decode()

    print('Sending email... SUCCESS\n\n')
    message = service.users().messages().send(userId = 'me', body={'raw': raw_string}).execute()

  


if __name__ == '__main__':
    main()